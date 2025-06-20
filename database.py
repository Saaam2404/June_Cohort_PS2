import datetime
import os
from openpyxl import Workbook, load_workbook

class Appointments:
    @staticmethod
    def save_appointment(patient_name,age,gender,address,symptoms, appointment_date, filename="appointments.xlsx"):
        # Tis  ensures the directory exists
        directory = os.path.dirname(filename)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)

        # Appointment details
        appointment_details = {
            "Patient Name": patient_name,
            "Age":age,
            "Gender":gender,
            "Address":address,
            "Symptoms": ", ".join(symptoms),
            "Appointment Date": appointment_date,
            "Timestamp": datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        }

        try:
            # Try to load the existing workbook
            wb = load_workbook(filename)
            sheet = wb.active
        except FileNotFoundError:
            # If the file doesn't exist, a new workbook is created
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Patient Name","Age","Gender","Address", "Symptoms","Appointment Date", "Timestamp"])

        # Appending the new appointment details
        sheet.append([appointment_details["Patient Name"],appointment_details["Age"] ,appointment_details["Gender"] ,appointment_details["Address"] ,appointment_details["Symptoms"], appointment_details["Appointment Date"], appointment_details["Timestamp"]])
        wb.save(filename)
        print("Appointment saved successfully.")

    @staticmethod
    def display_appointments(filename="appointments.xlsx"):
        try:
            # Loading the existing workbook
            wb = load_workbook(filename)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                print(row)
        except FileNotFoundError:
            print("No appointments found.")